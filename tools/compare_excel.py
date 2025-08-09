import sys
from pathlib import Path
from openpyxl import load_workbook


def summarize_workbook(path: Path):
    wb = load_workbook(path, data_only=True)
    info = {
        'file': str(path),
        'sheets': []
    }
    for ws in wb.worksheets:
        # Determine used range roughly
        max_row = ws.max_row
        max_col = ws.max_column
        # Sample first 5 rows and 10 columns
        sample = []
        for r in range(1, min(max_row, 5) + 1):
            row = []
            for c in range(1, min(max_col, 10) + 1):
                v = ws.cell(row=r, column=c).value
                row.append(v)
            sample.append(row)
        info['sheets'].append({
            'name': ws.title,
            'size': (max_row, max_col),
            'sample': sample
        })
    return info


def compare_sheets(info1, info2):
    # Build name map
    names1 = {s['name'] for s in info1['sheets']}
    names2 = {s['name'] for s in info2['sheets']}
    same = names1 & names2
    only1 = names1 - names2
    only2 = names2 - names1
    return same, only1, only2


def print_summary(info):
    print(f"File: {info['file']}")
    for s in info['sheets']:
        rows, cols = s['size']
        print(f"  - Sheet: {s['name']}  Size: {rows}x{cols}")
        for row in s['sample']:
            print("    ", row)
    print()


def main():
    if len(sys.argv) != 3:
        print("Usage: python compare_excel.py <original.xlsx> <translated.xlsx>")
        sys.exit(1)
    p1 = Path(sys.argv[1])
    p2 = Path(sys.argv[2])
    if not p1.exists() or not p2.exists():
        print("One of the files does not exist.")
        sys.exit(2)
    info1 = summarize_workbook(p1)
    info2 = summarize_workbook(p2)
    print("=== Original Summary ===")
    print_summary(info1)
    print("=== Translated Summary ===")
    print_summary(info2)
    same, only1, only2 = compare_sheets(info1, info2)
    print(f"Common sheets: {sorted(same)}")
    print(f"Only in original: {sorted(only1)}")
    print(f"Only in translated: {sorted(only2)}")


if __name__ == '__main__':
    main()

