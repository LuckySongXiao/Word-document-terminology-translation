#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel文档处理器
支持.xlsx和.xls格式的Excel文件翻译
支持术语占位符替换预处理翻译法
"""

import os
import logging
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment
from datetime import datetime
from typing import Dict, List, Any
import re
from utils.term_extractor import TermExtractor
from .translation_detector import TranslationDetector

logger = logging.getLogger(__name__)

class ExcelProcessor:
    """Excel文档处理器"""

    def __init__(self, translator):
        """
        初始化Excel处理器

        Args:
            translator: 翻译服务实例
        """
        self.translator = translator
        self.term_extractor = TermExtractor()
        self.translation_detector = TranslationDetector()  # 翻译检测器
        self.skip_translated_content = True  # 是否跳过已翻译内容（默认启用）
        self.source_lang = "zh"
        self.target_lang = "en"
        self.is_cn_to_foreign = True
        self.preprocess_terms = False
        self.reversed_terminology = {}
        self.progress_callback = None  # 进度回调函数

        # 配置日志记录器 - 添加Web日志记录器以确保日志同步
        self.web_logger = logging.getLogger('web_logger')

        # Excel特定的提示词
        self.excel_prompt = """你是一位专业的Excel文档翻译专家。请将以下文本翻译，并严格遵循以下要求：

1. 这是Excel表格中的内容，可能包含数据、标题、说明等
2. 保持原文的格式和结构
3. 对于数字、公式、单位等保持原样
4. 专业术语要准确翻译
5. 只返回翻译结果，不要添加解释或说明
6. 保持简洁和专业性"""

    def set_progress_callback(self, callback):
        """设置进度回调函数"""
        self.progress_callback = callback

    def _update_progress(self, progress: float, message: str = ""):
        """更新进度"""
        if self.progress_callback:
            import asyncio
            try:
                # 尝试异步调用回调函数
                if asyncio.iscoroutinefunction(self.progress_callback):
                    # 检查是否已有事件循环在运行
                    try:
                        loop = asyncio.get_event_loop()
                        if loop.is_running():
                            # 如果当前事件循环正在运行，创建一个任务而不是新的事件循环
                            asyncio.create_task(self.progress_callback(progress, message))
                        else:
                            # 如果当前事件循环不在运行，使用它来运行协程
                            loop.run_until_complete(self.progress_callback(progress, message))
                    except RuntimeError:
                        # 如果没有事件循环，创建一个新的
                        loop = asyncio.new_event_loop()
                        asyncio.set_event_loop(loop)
                        try:
                            loop.run_until_complete(self.progress_callback(progress, message))
                        finally:
                            loop.close()
            except Exception as e:
                logger.error(f"更新进度失败: {str(e)}")

        # 记录进度到日志
        if progress >= 0:
            logger.info(f"进度: {progress*100:.1f}% - {message}")
        else:
            logger.error(f"进度: 错误 - {message}")

    def process_document(self, file_path: str, target_language: str, terminology: Dict,
                        source_lang: str = "zh", target_lang: str = "en") -> str:
        """
        处理Excel文档翻译

        Args:
            file_path: 文件路径
            target_language: 目标语言名称（用于术语表查找）
            terminology: 术语词典
            source_lang: 源语言代码，默认为中文(zh)
            target_lang: 目标语言代码，默认为英文(en)

        Returns:
            str: 输出文件路径
        """
        # 更新进度：开始处理
        self._update_progress(0.01, "开始处理Excel文档...")

        # 设置翻译参数
        self.source_lang = source_lang
        self.target_lang = target_lang
        self.target_language = target_language  # 保存目标语言名称

        # 从实例属性获取输出格式和术语预处理设置（如果没有设置则使用默认值）
        output_format = getattr(self, 'output_format', 'bilingual')
        preprocess_terms = getattr(self, 'preprocess_terms', False)

        # 根据源语言和目标语言确定翻译方向
        self.is_cn_to_foreign = source_lang == "zh" or (source_lang == "auto" and target_lang != "zh")
        logger.info(f"翻译方向: {'中文→外语' if self.is_cn_to_foreign else '外语→中文'}")
        logger.info(f"输出格式: {'双语对照' if output_format == 'bilingual' else '仅翻译结果'}")
        logger.info(f"术语预处理: {'启用' if preprocess_terms else '禁用'}")

        # 检查文件是否存在
        if not os.path.exists(file_path):
            logger.error(f"文件不存在: {file_path}")
            raise Exception("文件不存在")

        # 创建输出目录
        output_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "输出")
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)

        # 更新进度：准备文件
        self._update_progress(0.05, "准备输出文件...")

        # 生成输出文件名
        file_name = os.path.splitext(os.path.basename(file_path))[0]
        time_stamp = datetime.now().strftime("%Y%m%d%H%M%S")
        ext = os.path.splitext(file_path)[1]
        output_path = os.path.join(output_dir, f"{file_name}_翻译版_{time_stamp}{ext}")

        # 翻译结果存储
        translation_results = []
        used_terminology = {}

        # 更新进度：处理术语预处理
        self._update_progress(0.1, "处理术语预处理...")

        # 处理术语预处理
        if preprocess_terms and terminology:
            logger.info("启用术语预处理功能")
            logger.info(f"术语库类型: {type(terminology)}")
            logger.info(f"目标语言: {target_language}")

            if not self.is_cn_to_foreign:
                # 外语→中文翻译，创建反向术语库缓存
                # 从术语库中提取目标语言的术语
                target_terms = terminology.get(target_language, {}) if isinstance(terminology, dict) else {}
                logger.info(f"目标语言术语库类型: {type(target_terms)}")
                logger.info(f"目标语言术语库大小: {len(target_terms) if target_terms else 0}")

                if target_terms:
                    self.reversed_terminology = {v: k for k, v in target_terms.items() if v and k}
                    logger.info(f"创建反向术语库缓存，包含 {len(self.reversed_terminology)} 个术语对")
                else:
                    logger.warning(f"目标语言 '{target_language}' 的术语库为空")
                    self.reversed_terminology = {}

        try:
            # 更新进度：读取Excel文件
            self._update_progress(0.15, "读取Excel文件...")

            # 读取Excel文件
            if ext.lower() == '.xlsx':
                workbook = openpyxl.load_workbook(file_path)
            else:
                # 对于.xls文件，先用pandas读取再转换
                df_dict = pd.read_excel(file_path, sheet_name=None, engine='xlrd')
                workbook = openpyxl.Workbook()
                workbook.remove(workbook.active)  # 移除默认工作表

                for sheet_name, df in df_dict.items():
                    ws = workbook.create_sheet(title=sheet_name)
                    for r_idx, row in enumerate(df.values, 1):
                        for c_idx, value in enumerate(row, 1):
                            if pd.notna(value):
                                ws.cell(row=r_idx, column=c_idx, value=str(value))

            logger.info(f"成功读取Excel文件: {file_path}")

            # 更新进度：开始处理工作表
            self._update_progress(0.2, "开始处理工作表...")

            # 处理每个工作表
            total_sheets = len(workbook.sheetnames)
            for i, sheet_name in enumerate(workbook.sheetnames):
                logger.info(f"正在处理工作表: {sheet_name}")
                worksheet = workbook[sheet_name]

                # 计算当前工作表的进度（20%-80%之间）
                sheet_progress = 0.2 + (i / total_sheets) * 0.6
                self._update_progress(sheet_progress, f"处理工作表: {sheet_name}")

                self._process_worksheet(worksheet, terminology, translation_results, used_terminology)

            # 更新进度：保存文件
            self._update_progress(0.85, "保存翻译后的文件...")

            # 保存翻译后的文件
            workbook.save(output_path)
            logger.info(f"翻译完成，文件已保存到: {output_path}")

            # 更新进度：导出结果
            self._update_progress(0.9, "导出翻译结果...")

            # 导出翻译结果到Excel
            if translation_results:
                self.export_to_excel(translation_results, file_path, target_language)

            # 如果使用了术语预处理，导出使用的术语
            if used_terminology:
                self.export_used_terminology(used_terminology, file_path)

            # 更新进度：完成
            self._update_progress(1.0, "Excel翻译完成！")

            return output_path

        except Exception as e:
            logger.error(f"处理Excel文件时出错: {str(e)}")
            raise Exception(f"处理Excel文件失败: {str(e)}")

    def _process_worksheet(self, worksheet: Any, terminology: Dict,
                          translation_results: List, used_terminology: Dict) -> None:
        """处理单个工作表"""
        # 获取目标语言的术语库
        target_language = getattr(self, 'target_language', '英语')
        target_terms = terminology.get(target_language, {}) if isinstance(terminology, dict) else {}

        # 遍历所有有内容的单元格
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    # 处理字符串类型的单元格
                    if isinstance(cell.value, str) and cell.value.strip():
                        original_text = cell.value.strip()

                        # 检查是否应该跳过翻译
                        if self._should_skip_cell(original_text):
                            # 跳过翻译但保持原始内容
                            logger.debug(f"跳过翻译单元格 {cell.coordinate}: {original_text}")
                            continue

                        # 提取术语（如果启用术语预处理）
                        preprocess_terms = getattr(self, 'preprocess_terms', False)
                        if preprocess_terms and target_terms:
                            if self.source_lang == "zh":
                                # 中文 → 外语
                                cell_terms = self.term_extractor.extract_terms(original_text, target_terms)
                            else:
                                # 外语 → 中文，使用缓存的反向术语库
                                if hasattr(self, 'reversed_terminology') and self.reversed_terminology:
                                    cell_terms = self.term_extractor.extract_foreign_terms_from_reversed_dict(
                                        original_text, self.reversed_terminology)
                                else:
                                    cell_terms = self.term_extractor.extract_foreign_terms_by_chinese_values(
                                        original_text, target_terms)

                            # 更新使用的术语词典
                            used_terminology.update(cell_terms)

                        # 翻译单元格内容
                        translated_text = self._translate_cell_content(
                            original_text, terminology, used_terminology)

                        # 根据输出格式设置单元格内容
                        output_format = getattr(self, 'output_format', 'bilingual')
                        if output_format == "translation_only":
                            # 仅翻译结果
                            cell.value = translated_text
                        else:
                            # 双语对照（默认）
                            cell.value = f"{original_text}\n{translated_text}"
                            # 设置单元格样式以支持换行
                            cell.alignment = Alignment(wrap_text=True, vertical='top')

                        # 记录翻译结果
                        translation_results.append({
                            "工作表": worksheet.title,
                            "位置": f"{cell.coordinate}",
                            "原文": original_text,
                            "译文": translated_text
                        })

                    # 对于非字符串类型的单元格（数字、日期等），保持原样
                    # 这些内容会自动保留在输出文件中

    def _should_skip_cell(self, text: str) -> bool:
        """判断是否应该跳过翻译的单元格"""
        # 如果禁用了跳过已翻译内容功能，只进行基本检查
        if not self.skip_translated_content:
            # 只跳过明显的非文本内容（数字、公式等）
            should_skip, reason = self.translation_detector.should_skip_translation(text, self.source_lang, self.target_lang)
            if should_skip and any(keyword in reason for keyword in ["纯数字", "代码", "URL", "邮箱", "公式"]):
                logger.debug(f"跳过Excel非文本内容: {reason} - {text[:30]}...")
                return True
            return False

        # 使用新的翻译检测器
        should_skip, reason = self.translation_detector.should_skip_translation(text, self.source_lang, self.target_lang)
        if should_skip:
            logger.debug(f"跳过Excel单元格翻译: {reason} - {text[:30]}...")
            return True

        # Excel特有的跳过条件
        # 跳过公式
        if text.startswith('='):
            logger.debug(f"跳过Excel公式: {text[:30]}...")
            return True

        return False

    def _translate_cell_content(self, text: str, terminology: Dict, used_terminology: Dict) -> str:
        """翻译单元格内容"""
        try:
            preprocess_terms = getattr(self, 'preprocess_terms', False)
            if preprocess_terms and used_terminology:
                # 使用术语预处理方式翻译
                if self.source_lang == "zh":
                    # 中文 → 外语
                    processed_text = self.term_extractor.replace_terms_with_placeholders(text, used_terminology)
                    translated_with_placeholders = self.translator.translate_text(
                        processed_text,
                        None,
                        self.source_lang,
                        self.target_lang,
                        prompt=self.excel_prompt
                    )
                    translation = self.term_extractor.restore_placeholders_with_foreign_terms(translated_with_placeholders)
                else:
                    # 外语 → 中文
                    reverse_terminology = used_terminology
                    processed_text = self.term_extractor.replace_foreign_terms_with_placeholders(text, reverse_terminology)
                    translated_with_placeholders = self.translator.translate_text(
                        processed_text,
                        None,
                        self.source_lang,
                        self.target_lang,
                        prompt=self.excel_prompt
                    )
                    translation = self.term_extractor.restore_placeholders_with_chinese_terms(translated_with_placeholders)
            else:
                # 使用常规方式翻译
                # 获取目标语言的术语库
                target_language = getattr(self, 'target_language', '英语')
                target_terms = terminology.get(target_language, {}) if isinstance(terminology, dict) else {}

                translation = self.translator.translate_text(
                    text,
                    target_terms if target_terms else None,
                    self.source_lang,
                    self.target_lang,
                    prompt=self.excel_prompt
                )

            return translation.strip()

        except Exception as e:
            logger.error(f"翻译单元格内容失败: {str(e)}")
            return f"[翻译失败: {str(e)}]"

    def export_to_excel(self, translation_results: List[Dict], file_path: str, target_language: str = None) -> None:
        """将翻译结果导出到Excel文件"""
        try:
            # 创建输出目录
            output_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "输出")
            if not os.path.exists(output_dir):
                os.makedirs(output_dir)

            # 获取文件名
            file_name = os.path.splitext(os.path.basename(file_path))[0]
            time_stamp = datetime.now().strftime("%Y%m%d%H%M%S")

            # 根据目标语言生成文件名
            lang_suffix = f"_{target_language}" if target_language else ""
            excel_path = os.path.join(output_dir, f"{file_name}_翻译结果{lang_suffix}_{time_stamp}.xlsx")

            # 转换翻译结果为DataFrame
            df = pd.DataFrame(translation_results)

            # 保存到Excel
            df.to_excel(excel_path, index=False, engine='openpyxl')
            logger.info(f"翻译结果已导出到: {excel_path}")

        except Exception as e:
            logger.error(f"导出翻译结果失败: {str(e)}")

    def export_used_terminology(self, used_terminology: Dict, file_path: str) -> None:
        """导出使用的术语"""
        try:
            # 创建输出目录
            output_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "输出")
            if not os.path.exists(output_dir):
                os.makedirs(output_dir)

            # 获取文件名
            file_name = os.path.splitext(os.path.basename(file_path))[0]
            time_stamp = datetime.now().strftime("%Y%m%d%H%M%S")
            excel_path = os.path.join(output_dir, f"{file_name}_使用术语_{time_stamp}.xlsx")

            # 转换术语为DataFrame
            if self.is_cn_to_foreign:
                df = pd.DataFrame(list(used_terminology.items()), columns=['中文术语', '外语术语'])
            else:
                df = pd.DataFrame(list(used_terminology.items()), columns=['外语术语', '中文术语'])

            # 保存到Excel
            df.to_excel(excel_path, index=False, engine='openpyxl')
            logger.info(f"使用的术语已导出到: {excel_path}")

        except Exception as e:
            logger.error(f"导出使用术语失败: {str(e)}")
